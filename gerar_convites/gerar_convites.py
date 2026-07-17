from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont, TTFError
from reportlab.lib.colors import HexColor
from reportlab.lib.utils import ImageReader
from openpyxl import load_workbook
from PyPDF2 import PdfReader, PdfWriter
import io
import os
import sys
import qrcode
from urllib.parse import quote

# Configurar saída do console para UTF-8 (evita erros com emojis no Windows)
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

# =================================================================================
# ⚙️ CONFIGURAÇÕES
# =================================================================================
# ❗️❗️❗️ IMPORTANTE: Verifique se esta é a URL base correta do seu site!
URL_BASE_DO_SITE = "https://casamento-evellyn-e-guilherme.netlify.app"
# =================================================================================

os.makedirs("convites", exist_ok=True)

# 🔹 Registrar fontes
nome_fonte_cursiva = "FonteCursiva"

# Para usar uma fonte nova e mais desenhada (cursiva):
# 1. Baixe o arquivo da fonte (ex: .ttf) e coloque nesta pasta 'gerar_convites'.
#    Sugestão: "Dancing Script" do Google Fonts.
try:
    # Tenta registrar a fonte cursiva personalizada que dá um toque especial
    pdfmetrics.registerFont(TTFont("FonteCursiva", "DancingScript-Regular.ttf"))
except TTFError:
    # Se não encontrar DancingScript, tenta byrani.ttf que está na pasta
    try:
        pdfmetrics.registerFont(TTFont("FonteCursiva", "byrani.ttf"))
        print("\nℹ️  INFO: Usando a fonte 'byrani.ttf' para os nomes.\n")
    except TTFError:
        # Se nenhuma das fontes cursivas .ttf for encontrada, usa Helvetica-Bold (padrão do PDF)
        print("\n⚠️  AVISO: Nenhuma fonte cursiva personalizada (.ttf) foi encontrada.")
        print("   -> Usando a fonte padrão 'Helvetica-Bold' para os nomes.\n")
        nome_fonte_cursiva = "Helvetica-Bold"

pdfmetrics.registerFont(TTFont('Vera', 'Vera.ttf')) # Fonte padrão para o texto do QR Code

# Carregar planilha do Excel
wb = load_workbook("convidados.xlsx")
sheet = wb.active

# Obter dimensões do template
reader_temp = PdfReader("Convite_Template.pdf")
largura, altura = float(reader_temp.pages[0].mediabox.width), float(reader_temp.pages[0].mediabox.height)

print("🚀 Iniciando a geração dos convites com QR Code personalizado...")

for linha in sheet.iter_rows(min_row=2, values_only=True):
    nome = linha[0]
    if not nome:
        continue

    print(f"  -> Gerando para: {nome}")

    # --- Gera um QR Code personalizado para cada convidado ---
    url_personalizada = f"{URL_BASE_DO_SITE}/pages/invite.html?name={quote(nome)}"
    qr_img = qrcode.make(url_personalizada)
    
    qr_img_bytes = io.BytesIO()
    qr_img.save(qr_img_bytes, format='PNG')
    qr_img_bytes.seek(0)
    qr_code_para_pdf = ImageReader(qr_img_bytes)
    # ----------------------------------------------------

    packet = io.BytesIO()
    c = canvas.Canvas(packet, pagesize=(largura, altura))

    # 🔳 Adiciona o QR Code e o texto
    qr_size = 80
    left_margin = 20  # Margem da esquerda para o alinhamento
    posicao_y_qr = 10
    
    c.setFont("Vera", 9)
    c.setFillColor(HexColor("#333333"))
    # Desenha o QR Code na posição esquerda
    c.drawImage(qr_code_para_pdf, left_margin, posicao_y_qr, width=qr_size, height=qr_size, mask='auto')

    # Adiciona um link clicável sobre o QR Code
    link_rect = (left_margin, posicao_y_qr, left_margin + qr_size, posicao_y_qr + qr_size)
    c.linkURL(url_personalizada, link_rect, relative=1)

    # Adiciona o texto informativo ao lado do QR Code
    text_x = left_margin + qr_size + 10  # Posição X do texto
    text_y_start = posicao_y_qr + (qr_size / 2) + 10 # Começa um pouco acima do meio
    c.drawString(text_x, text_y_start, "QR code para o seu")
    c.drawString(text_x, text_y_start - 12, "convite virtual e")
    c.drawString(text_x, text_y_start - 24, "lista de presentes.")
    
    c.save()
    packet.seek(0)

    # Mistura o conteúdo gerado com o template PDF
    template = PdfReader("Convite_Template.pdf")
    pagina_base = template.pages[0]

    overlay = PdfReader(packet)
    pagina_base.merge_page(overlay.pages[0])

    writer = PdfWriter()
    writer.add_page(pagina_base)

    # Adiciona as demais páginas do template sem alterações (ex: a 2ª página)
    for i in range(1, len(template.pages)):
        writer.add_page(template.pages[i])

    # Salva o arquivo final
    nome_arquivo = "".join(c for c in nome if c.isalnum() or c in (' ',)).rstrip()
    with open(f"convites/convite_{nome_arquivo.replace(' ', '_')}.pdf", "wb") as f:
        writer.write(f)

print("\n✅ Convites gerados com sucesso na pasta 'convites'!")
